"""
Load colony data from Excel file using only openpyxl (no pandas).
Prefers SummaryFileGenerated.xlsx (has real coordinates), falls back to Colibri file.
SummaryFileGenerated columns: Year, Date, State, GeoRegion, ColonyName, SpeciesCode, Longitude_y, Latitude_y, Nests, Birds
Colibri columns: Year, Date, State, GeoRegion, ColonyName, SpeciesCode, Nests, Birds, CombinedMayJuneTotal?
"""
import hashlib
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook

# Approximate coastal bounds by state (lat, lon) for placing colonies
STATE_BOUNDS = {
    "LA": {"lat_min": 28.8, "lat_max": 30.2, "lon_min": -93.5, "lon_max": -89.0},
    "TX": {"lat_min": 26.0, "lat_max": 29.8, "lon_min": -97.5, "lon_max": -93.5},
    "FL": {"lat_min": 24.5, "lat_max": 30.5, "lon_min": -87.5, "lon_max": -80.0},
    "MS": {"lat_min": 30.0, "lat_max": 30.5, "lon_min": -89.5, "lon_max": -88.2},
    "AL": {"lat_min": 30.0, "lat_max": 30.5, "lon_min": -88.2, "lon_max": -87.3},
}

SPECIES_CODES = {
    "BLSK": "Black Skimmer",
    "BRPE": "Brown Pelican",
    "LAGU": "Laughing Gull",
    "ROYT": "Royal Tern",
    "SATE": "Sandwich Tern",
    "CATE": "Caspian Tern",
    "GBTE": "Great Black-backed Tern",
    "RUTU": "Ruddy Turnstone",
    "AMOY": "American Oystercatcher",
    "FOTE": "Forster's Tern",
    "BCNH": "Black-crowned Night-Heron",
    "CAEG": "Great Egret",
    "SNEG": "Snowy Egret",
    "TRHE": "Tricolored Heron",
    "WHIB": "White Ibis",
    "ROSP": "Roseate Spoonbill",
    "GBHE": "Great Blue Heron",
    "GREG": "Great Egret",
    "REEG": "Reddish Egret",
    "LBHE": "Little Blue Heron",
    "DCCO": "Double-crested Cormorant",
    "OSPR": "Osprey",
    "AWPE": "American White Pelican",
    "BNST": "Black-necked Stilt",
    "WIPH": "Wilson's Plover",
    "ANHI": "Anhinga",
    "AMCO": "American Coot",
    "NECO": "Neotropic Cormorant",
    "AMAV": "American Avocet",
    "CANG": "Canada Goose",
    "GLIB": "Glossy Ibis",
    "LBBG": "Lesser Black-backed Gull",
    "NSHO": "Northern Shoveler",
    "UNWA": "Unidentified Waterbird",
    "BBWD": "Black-bellied Whistling-Duck",
    "SOTE": "Snowy Egret",
    "COGA": "Common Gallinule",
    "GTGR": "Great-tailed Grackle",
    "WFIB": "White-faced Ibis",
    "SDHE": "Snowy Egret",
    "WHEG": "White Egret",
    "BBPL": "Black-bellied Plover",
    "REKN": "Red Knot",
    "UNCO": "Unidentified Cormorant",
    "MAGO": "Marbled Godwit",
    "SBDO": "Short-billed Dowitcher",
    "GRFL": "Great Blue Heron",
    "FICR": "Fish Crow",
    "CRCA": "Crested Caracara",
    "BWTE": "Blue-winged Teal",
    "WILL": "Willet",
    "FUWD": "Fulvous Whistling-Duck",
    "HERG": "Herring Gull",
    "UNSB": "Unidentified Seabird",
    "LETE": "Least Tern",
}


def _colony_coords(state: str, geo_region: str, colony_name: str) -> tuple:
    """Deterministic lat/lon from colony identity (no coords in Excel)."""
    key = f"{state}|{geo_region}|{colony_name}"
    h = hashlib.sha256(key.encode()).hexdigest()
    lat_norm = int(h[:8], 16) / (16 ** 8)
    lon_norm = int(h[8:16], 16) / (16 ** 8)
    bounds = STATE_BOUNDS.get(state, STATE_BOUNDS["LA"])
    lat = bounds["lat_min"] + (bounds["lat_max"] - bounds["lat_min"]) * lat_norm
    lon = bounds["lon_min"] + (bounds["lon_max"] - bounds["lon_min"]) * lon_norm
    return round(lat, 5), round(lon, 5)


def _col_index(header_row: tuple, want: str) -> int:
    """Get 0-based column index. want is one of: year, nests, state, geo_region, colony_name, species_code, combined, longitude, latitude."""
    want = want.lower().replace(" ", "")
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "")
        if want == "year" and c == "year":
            return i
        if want == "nests" and c == "nests":
            return i
        if want == "state" and c == "state":
            return i
        if want == "geo_region" and (c == "georegion" or "geo" in c):
            return i
        if want == "colony_name" and (c == "colonyname" or "colony" in c):
            return i
        if want == "species_code" and (c == "speciescode" or "species" in c):
            return i
        if want == "combined" and "combined" in c:
            return i
        if want == "longitude" and ("longitude" in c or c == "lon"):
            return i
        if want == "latitude" and ("latitude" in c or c == "lat"):
            return i
    return -1


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v, default: int = 0) -> int:
    if v is None:
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _float(v, default: float = 0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def load_colony_data_from_excel(file_path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """
    Load colony totals from Excel (openpyxl only, no pandas).
    Prefers SummaryFileGenerated.xlsx (has real coordinates), falls back to Colibri.
    Returns list of records: colony_id, site_index, year, species, nest_count, latitude, longitude.
    """
    if file_path is None:
        project_root = Path(__file__).resolve().parent.parent.parent.parent
        # Try SummaryFileGenerated first (has real coordinates)
        summary_file = project_root / "SummaryFileGenerated.xlsx"
        colibri_file = project_root / "Colibri2010-21ColonyTotalsMayJuneCombined_8Nov22.xlsx"
        if summary_file.exists():
            file_path = summary_file
        elif colibri_file.exists():
            file_path = colibri_file
        else:
            return []
    
    if not file_path.exists():
        return []

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    header = rows[0]
    idx_year = _col_index(header, "year")
    idx_nests = _col_index(header, "nests")
    idx_state = _col_index(header, "state")
    idx_geo = _col_index(header, "geo_region")
    idx_colony = _col_index(header, "colony_name")
    idx_species = _col_index(header, "species_code")
    idx_combined = _col_index(header, "combined")
    idx_lon = _col_index(header, "longitude")
    idx_lat = _col_index(header, "latitude")

    if idx_year < 0 or idx_nests < 0:
        return []

    has_real_coords = idx_lon >= 0 and idx_lat >= 0

    # Build unique colonies and coords
    seen = set()
    colony_list = []
    colony_coords = {}  # (state, geo, name) -> (lat, lon)
    
    for row in rows[1:]:
        if len(row) <= max(idx_year, idx_nests, idx_state, idx_geo, idx_colony):
            continue
        state = _norm(row[idx_state]) if idx_state >= 0 else ""
        geo = _norm(row[idx_geo]) if idx_geo >= 0 else ""
        name = _norm(row[idx_colony]) if idx_colony >= 0 else ""
        key = (state, geo, name)
        if key in seen:
            continue
        seen.add(key)
        colony_list.append(key)
        
        # Store real coordinates if available
        if has_real_coords:
            lon_val = _float(row[idx_lon] if idx_lon >= 0 and len(row) > idx_lon else None, None)
            lat_val = _float(row[idx_lat] if idx_lat >= 0 and len(row) > idx_lat else None, None)
            if lon_val is not None and lat_val is not None:
                if -180 <= lon_val <= 180 and -90 <= lat_val <= 90:
                    colony_coords[key] = (round(lat_val, 5), round(lon_val, 5))

    colony_list.sort()
    colony_lookup = {}
    for site_index, (state, geo, name) in enumerate(colony_list):
        key = (state, geo, name)
        if key in colony_coords:
            lat, lon = colony_coords[key]
        else:
            lat, lon = _colony_coords(state, geo, name)
        colony_lookup[key] = {"site_index": site_index, "latitude": lat, "longitude": lon}

    records = []
    for row in rows[1:]:
        if len(row) <= max(idx_year, idx_nests, idx_state, idx_geo, idx_colony, idx_species):
            continue
        year = _num(row[idx_year])
        nests = _num(row[idx_nests])
        if year < 2010 or year > 2022 or nests <= 0:
            continue
        state = _norm(row[idx_state]) if idx_state >= 0 else ""
        geo = _norm(row[idx_geo]) if idx_geo >= 0 else ""
        name = _norm(row[idx_colony]) if idx_colony >= 0 else ""
        key = (state, geo, name)
        if key not in colony_lookup:
            continue
        info = colony_lookup[key]
        code = _norm(row[idx_species]) if idx_species >= 0 else ""
        species = SPECIES_CODES.get(code.upper(), code or "Unidentified")
        colony_id = f"{state}-{name}"[:50].replace(" ", "-")
        records.append({
            "colony_id": colony_id,
            "site_index": info["site_index"],
            "year": year,
            "species": species,
            "nest_count": nests,
            "latitude": info["latitude"],
            "longitude": info["longitude"],
        })
    return records
